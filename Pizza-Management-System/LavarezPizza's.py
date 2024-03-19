from tkinter import *
import time
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import json

excel_con = Workbook()

main = tk.Tk()
main.geometry("1360x750")
main.title("PIZZA SAFARI")
main.resizable(False, False)
main.iconbitmap('p.ico')

cartlist=[]
amount=0

main_frame=Frame(main,height=150,width=1366)
main_frame.pack(fill=BOTH,expand=1)

mainlogo=PhotoImage(file="logo1.png")

mainl=Label(main_frame,image=mainlogo)
mainl.place(x=0,y=0)

main_frame2=Frame(main,height=618,width=1366)
main_frame2.pack(fill=BOTH,expand=1)

mainc=Canvas(main_frame2,height=618,width=1366)
mainc.pack()

mainback=PhotoImage(file="pizzamain.png")
mainc.create_image(680,284,image=mainback)

mainlab=Button(main_frame2,text= "Click Here to enter the World of Pizzas",command=lambda:mainlogin(),cursor="hand2", bd=10 ,font=("cooper black",30, 'bold'),fg="white",bg="#0b1335")
mainlab.place(x=250,y=250)

def mainlogin():
    login= Toplevel()
    login.title("Lavarez's Pizza")
    login.geometry("1366x768")
    login.resizable(False, False)
    main.withdraw()

    excel_con = load_workbook('Orders.xlsx')
    excel_activate = excel_con.active
    
    def home_function():
        login.destroy()
        main.deiconify()

    def login_func():
            excel_con = load_workbook('Account.xlsx')
            excel_activate = excel_con.active
            user1 = user.get()
            passw = pas.get()
            found = False

            if user1 == "" or passw == "":
                messagebox.showinfo("Login","Empty Entry is not allowed")
            else:
                for each_cell in range(2, excel_activate.max_row + 1):
                    if user1 == excel_activate['A' + str(each_cell)].value and passw == excel_activate['B' + str(each_cell)].value:
                        found = True
                        break
                if found:
                    excel_con = load_workbook('Orders.xlsx')
                    excel_activate = excel_con.active
                    messagebox.showinfo("Login","You have Successfully Log In\nWelcome to the Lavarez's Pizza")
                    welcome()
                else:
                    messagebox.showinfo("Login","You are Not Registered Yet")

    def Register():
        reg=Toplevel()
        reg.title("Lavarez's Pizza")
        reg.geometry("1366x768")
        reg.resizable(False, False)
        login.withdraw()

        excel_con = load_workbook('Account.xlsx')
        excel_activate = excel_con.active

        def back_to_login():
            reg.destroy()
            login.deiconify()
        
        def reg_func():
            Found = False
            user = usern.get()
            password = passd.get()
            em = email.get()
            address = last.get()
            firstname = first.get()
            mobileno = mob.get()
            
            if user == "" and password == "":
                messagebox.showinfo("ERROR", "FILL ALL ENTRIES")
            else:
                for each_cell in range(2, excel_activate.max_row + 1):
                    if user == excel_activate['A' + str(each_cell)].value:
                        Found = True
                        break
                    else:
                        Found = False
                if Found:
                    messagebox.showerror("ERROR", "Account Exist")
                else:
                    lastrow = str(excel_activate.max_row + 1)
                    excel_activate['A' + lastrow] = user
                    excel_activate['B' + lastrow] = password
                    excel_activate['C' + lastrow] = firstname
                    excel_activate['D' + lastrow] = address
                    excel_activate['E' + lastrow] = em
                    excel_activate['F' + lastrow] = mobileno
                    excel_con.save('Account.xlsx')
                    messagebox.showinfo("SUCCESS", "Account Created")
                    reg.destroy()
                    login.deiconify()

        regf1=Frame(reg,height=150,width=1366)
        regf1.pack(fill=BOTH,expand=1)

        logo=PhotoImage(file="logo.png")
        ba=Label(regf1,image=logo,height=150)
        ba.image = logo
        ba.place(x=0,y=0)

        home=Button(regf1,text="Home",command=lambda:home_function(),bg="#0b1335",cursor="hand2",fg="white",font=("default",16))
        home.place(x=925,y=100)

        localtime=time.asctime(time.localtime(time.time()))
        tim=Label(regf1,text=localtime,fg="white",font=("default",16),bg="#0b1335")
        tim.place(x=925,y=50)
        
        
        regf2=Frame(reg,height=618,width=1366)
        regf2.pack(fill=BOTH,expand=1)

        c=Canvas(regf2,height=618,width=1366)
        c.pack()
        logo1=PhotoImage(file="pizzamain.png")
        c.create_image(683,309,image=logo1)
        c.create_rectangle(150,100,1216,450,fill="#d3ede6",outline="white",width=6)

        log=Label(regf2,text="REGISTRATION",fg="white",bg="#0b1335",width=20,font=("cooper black",27))
        log.place(x=480,y=120)

        lab1=Label(regf2,text="FullName",bg="#d3ede6",font=("cooper black",18))
        lab1.place(x=190,y=200)

        first=Entry(regf2,bg="white",width=15,font=("cooper black",18),bd=5)
        first.place(x=430,y=200)

        lab2=Label(regf2,text="Address",bg="#d3ede6",font=("cooper black",18))
        lab2.place(x=730,y=200)

        last=Entry(regf2,bg="white",width=15,font=("cooper black",18),bd=5)
        last.place(x=920,y=200)

        lab3=Label(regf2,text="Username",bg="#d3ede6",font=("cooper black",18))
        lab3.place(x=190,y=250)

        usern=Entry(regf2,bg="white",width=15,font=("cooper black",18),bd=5)
        usern.place(x=430,y=250)

        lab4=Label(regf2,text="Password",bg="#d3ede6",font=("cooper black",18))
        lab4.place(x=730,y=250)

        passd=Entry(regf2,bg="white",width=15,font=("cooper black",18),bd=5,show="•")
        passd.place(x=920,y=250)

        lab5=Label(regf2,text="Email",bg="#d3ede6",font=("cooper black",18))
        lab5.place(x=190,y=300)

        email=Entry(regf2,bg="white",width=15,font=("cooper black",18),bd=5)
        email.place(x=430,y=300)

        lab6=Label(regf2,text="Mobile No.",bg="#d3ede6",font=("cooper black",18))
        lab6.place(x=730,y=300)

        mob=Entry(regf2,bg="white",width=15,font=("cooper black",18),bd=5)
        mob.place(x=920,y=300)

        bc=Button(regf2,text="Back",cursor="hand2",command=lambda:back_to_login(),fg="white",bg="#0b1335",font=("cooper black",18),bd=5)
        bc.place(x=370,y=370)

        rg=Button(regf2,text="Register",cursor="hand2",fg="white",bg="#0b1335",command=lambda:reg_func(),font=("cooper black",18),bd=5)
        rg.place(x=610,y=370)
        
        def clear():
            usern.delete(0,END)
            passd.delete(0,END)
            first.delete(0,END)
            last.delete(0,END)
            email.delete(0,END)
            mob.delete(0,END)

        cl=Button(regf2,text="Clear",cursor="hand2",fg="white",bg="#0b1335",command=lambda:clear(),font=("cooper black",18),bd=5)
        cl.place(x=910,y=370)
        
        reg.mainloop()
    

    loginf1=Frame(login,height=150,width=1366)
    loginf1.pack(fill=BOTH,expand=1)
    logo=PhotoImage(file="logo.PNG")
    logo_lbl=Label(loginf1,image=logo,height=150)
    logo_lbl.image = logo
    logo_lbl.place(x=0,y=0)

    home=Button(loginf1,text="Home",command=lambda:home_function(),bg="#0b1335",cursor="hand2",bd=4,fg="white",font=("cooper black",16))
    home.place(x=925,y=100)

    localtime=time.asctime(time.localtime(time.time()))
    tim=Label(loginf1,text=localtime,fg="white",font=("default",16),bg="#0b1335")
    tim.place(x=925,y=50)

    loginf2=Frame(login,height=618,width=1366)
    loginf2.pack(fill=BOTH,expand=1)

    c=Canvas(loginf2,height=618,width=1366)
    c.pack()
    logo1=PhotoImage(file="pizzamain.png")
    c.image = logo1
    c.create_image(683,309,image=logo1)
    c.create_rectangle(50,100,700,450,fill="#d3ede6",outline="white",width=6)

    log=Label(loginf2,text="LOGIN",fg="white",bg="#0b1335",width=26,font=("cooper black",27))
    log.place(x=59,y=105)

    lab1=Label(loginf2,text="UserName",bg="#d3ede6",font=("cooper black",22))
    lab1.place(x=100,y=180)

    user=Entry(loginf2,bg="white",font=("cooper black",22),bd=6 ,justify='left')
    user.place(x=320,y=180)

    lab2=Label(loginf2,text="Password",bg="#d3ede6",font=("cooper black",22))
    lab2.place(x=105,y=250)

    pas=Entry(loginf2,bg="white",font=("cooper black",22),bd=6 ,justify='left',show="•")
    pas.place(x=320,y=250)

    lg=Button(loginf2,text="Login",cursor="hand2",command=lambda:login_func(),fg="white",bg="#0b1335",font=("cooper black",20),bd=4)
    lg.place(x=180,y=320)

    cl=Button(loginf2,text="Clear",cursor="hand2",command=lambda:clear(),fg="white",bg="#0b1335",font=("cooper black",20),bd=4)
    cl.place(x=450,y=320)

    rg=Button(loginf2,text="New to Lavarez's Pizza",command=lambda:Register(),fg="white",cursor="hand2",bg="#8c68c1",font=("cooper black",20),bd=6)
    rg.place(x=200,y=390)

    c.create_rectangle(850,120,1310,480,fill="#d3ede6",outline="white",width=4)
    ext=PhotoImage(file="p4.png")

    url=Label(loginf2,image=ext,cursor="hand2")
    url.image = ext
    url.place(x=855,y=125)

    def clear():
        global delete_user
        global delete_pas
        delete_user = user.delete(0,END)
        delete_pas = pas.delete(0,END)
        delete_user
        delete_pas
        
        login.mainloop()

    def welcome():
        welc=Toplevel()
        welc.title("Lavarez's Pizza")
        welc.geometry("1366x768")
        welc.resizable(False, False)
        login.withdraw()

        excel_con = load_workbook('Orders.xlsx')
        excel_activate = excel_con.active

        def back_to_login():
            welc.withdraw()
            login.deiconify()
            user.delete(0,END)
            pas.delete(0,END)


        pizf1=Frame(welc,height=150,width=1366)
        pizf1.pack(fill=BOTH,expand=1)

        c=Canvas(pizf1,height=150,width=1366)
        c.pack()
        logo=PhotoImage(file="logo.PNG")
        c.create_image(683,75,image=logo)
        c.image = logo
        c.create_text(950,80,text="WELCOME",fill="white",font=("default",20))
        name= user.get()
        c.create_text(950,120,text=name,fill="white",font=("default",18))

        out=Button(pizf1,text="Log Out",command=lambda:back_to_login(),bg="#0b1335",cursor="hand2",fg="white",font=("default",16))
        out.place(x=1200,y=100)

        localtime=time.asctime(time.localtime(time.time()))
        c.create_text(1000,40,text=localtime,fill="white",font=("default",16))
        
        pizf2=Frame(welc,height=618,width=1366)
        pizf2.pack(fill=BOTH,expand=1)

        c=Canvas(pizf2,height=618,width=1366)
        c.pack()
        logo1=PhotoImage(file="pizzamain.png")
        c.create_image(683,309,image=logo1)
        c.image = logo1
        c.create_rectangle(100,120,390,470,fill="#d3ede6",outline="white",width=2)

        deli=PhotoImage(file="delivery.png")
        c.image = deli
        c.create_image(240,260,image=deli)

        de=Button(pizf2,text="ORDER",cursor="hand2",fg="white",command=lambda:menulist(),bg="#0b1335",font=("default",20),bd=5)
        de.place(x=180,y=400)

        delete_button = Button(pizf2, text="Delete Orders",cursor="hand2",fg="white", command=lambda:delete_selected_item(Event),bg="#0b1335",font=("default",20),bd=5)
        delete_button.place(x=135,y=500)

        def refresh_data(tree):
            tree.delete(*tree.get_children())
            data = get_updated_data()
            for item in data:
                tree.insert("", "end", values=item)

        def get_updated_data():
            excel_con = load_workbook('Orders.xlsx')
            excel_activate = excel_con.active
            updated_value = list()
            for each_cell in range(2, (excel_activate.max_row) + 1):
                updated_value.append([excel_activate['A' + str(each_cell)].value,
                                    excel_activate['B' + str(each_cell)].value,
                                    excel_activate['C' + str(each_cell)].value,
                                    excel_activate['D' + str(each_cell)].value,
                                    excel_activate['E' + str(each_cell)].value,
                                    excel_activate['F' + str(each_cell)].value,
                                    excel_activate['G' + str(each_cell)].value])
            return updated_value
        

        treeframe = Frame(pizf2,width=850,height=500)
        treeframe.place(x=500,y=50)
        
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview", background="#323232", foreground="white", fieldbackground="#0b1335")
        global tree
        tree = ttk.Treeview(treeframe,height=25)
        tree['columns'] = ('FullName','Address','Email','Mobile No#','Username','Order','Total')

        tree.column('#0',width=0,stretch=0)
        tree.column('FullName',anchor='center',width=100)
        tree.column('Address',anchor='center',width=100)
        tree.column('Email',anchor='center',width=100)
        tree.column('Mobile No#',anchor='center',width=100)
        tree.column('Username',anchor='center',width=100)
        tree.column('Order',anchor='center',width=220)
        tree.column('Total',anchor='center',width=80)

        tree.heading('FullName',text='FullName')
        tree.heading('Address',text='Address')
        tree.heading('Email',text='Email')
        tree.heading('Mobile No#',text='Mobile No#')
        tree.heading('Username',text='Address')
        tree.heading('Order',text='Order')
        tree.heading('Total',text='Total')

        for each_cell in range(2, (excel_activate.max_row)+1):
            tree.insert(parent='', index="end", values=(excel_activate['A'+str(each_cell)].value,
                                                        excel_activate['B'+str(each_cell)].value, 
                                                        excel_activate['C'+str(each_cell)].value,
                                                        excel_activate['D'+str(each_cell)].value,
                                                        excel_activate['E'+str(each_cell)].value,
                                                        excel_activate['F'+str(each_cell)].value,
                                                        excel_activate['G'+str(each_cell)].value))
        tree.pack(fill='x')

        c.create_rectangle(105,125,378,465,outline="black",width=2)

        def delete_selected_item(event):
            selected_item = tree.focus()
            if selected_item:
                cell_address = each_cell
                tree.delete(selected_item)
                excel_activate.delete_rows(cell_address)

            excel_con.save('Orders.xlsx')
            return

        tree.bind('<Double-1>', lambda event: delete_selected_item())
        
        def menulist():
            menu=Toplevel()
            menu.title("Lavarez's Pizza")
            menu.geometry("1366x768")
            menu.resizable(False, False)
            welc.withdraw()
            cartlist = []

            def back_to_login():
                menu.destroy()
                login.deiconify()

            menuf1=Frame(menu,height=150,width=1366)
            menuf1.pack(fill=BOTH,expand=1)

            c=Canvas(menuf1,height=150,width=1366)
            c.pack()
            logo=PhotoImage(file="logo.PNG")
            c.image = logo
            c.create_image(683,75,image=logo)

            home=Button(menuf1,text="Log Out",command=lambda:back_to_login(),bg="#0b1335",cursor="hand2",fg="white",bd=5,font=("default",16,'bold'))
            home.place(x=1000,y=90)

            localtime=time.asctime(time.localtime(time.time()))
            c.create_text(1000,50,text=localtime,fill="white",font=("default",16))
            
            menuf2=Frame(menu,height=618,width=1366)
            menuf2.pack(fill=BOTH,expand=1)

            c=Canvas(menuf2,height=618,width=1366)
            c.pack()
            logo1=PhotoImage(file="pizzamain.png")
            c.image = logo1
            c.create_image(683,309,image=logo1)
            c.create_rectangle(50, 140, 1316, 420,fill="#d3ede6",outline="white",width=6)
            veg=PhotoImage(file="veg.png")
            c.image = veg
            c.create_image(230,250,image=veg)
            vegbut=Button(menuf2,text="Veg Pizza",cursor="hand2",fg="white",command=lambda:vegpizza(),bg="#0b1335",bd=5,font=("default",18,'bold'))
            vegbut.place(x=170,y=350)

            nonveg=PhotoImage(file="Non.png")
            c.image = nonveg
            c.create_image(530,250,image=nonveg)
            nonvegbut=Button(menuf2,text="Non-Veg Pizza",cursor="hand2",fg="white",command=lambda:nonvegpiz(),bg="#0b1335",bd=5,font=("default",18,'bold'))
            nonvegbut.place(x=440,y=350)

            chi=PhotoImage(file="chiken.png")
            c.image=chi
            c.create_image(830,250,image=chi)
            chibut=Button(menuf2,text="Special Chicken",cursor="hand2",fg="white",command=lambda:SpecialChi(),bg="#0b1335",bd=5,font=("default",18,'bold'))
            chibut.place(x=730,y=350)

            side=PhotoImage(file="extra.png")
            c.image = side
            c.create_image(1130,250,image=side)
            sidebut=Button(menuf2,text="Sides and Beverages",cursor="hand2",fg="white",command=lambda:sidebev(),bg="#0b1335",bd=5,font=("default",18,'bold'))
            sidebut.place(x=1000,y=350)

            amount = 0

            def addlist(q):
                global amount
                print("q:", q)
                if q[-2] != "0" and q[-2].isdigit():
                    cartlist.append(q)
                    amount += q[-1]
                    messagebox.showinfo("Cart","Item Successfully added")
                else:
                    messagebox.showinfo("Cart","Enter Valid Quantity to add")
                print(cartlist,amount)

            def Orderde():
                global amount  # Declare that you want to use the global 'amount' variable
                order=Toplevel()
                order.title("Lavarez's Pizza")
                order.geometry("1366x768")
                order.resizable(False, False)
                order.iconbitmap('p.ico')

                def back_to_login():
                    order.destroy()
                    login.deiconify()

                def addmore_function():
                    order.destroy()
                    menu.deiconify()

                def orderpay():
                    payorder = Toplevel()
                    payorder.geometry('1366x768')
                    payorder.title("Lavarez's Pizza")
                    payorder.resizable(False,False)
                    payorder.iconbitmap('p.ico')
                    order.withdraw()

                    excel_con = load_workbook('Orders.xlsx')
                    excel_activate = excel_con.active

                    def refresh_data(tree):
                        tree.delete(*tree.get_children())
                        data = get_updated_data()
                        for item in data:
                            tree.insert("", "end", values=item)

                    def get_updated_data():
                        excel_con = load_workbook('Orders.xlsx')
                        excel_activate = excel_con.active
                        updated_value = list()
                        for each_cell in range(2, (excel_activate.max_row) + 1):
                            updated_value.append([excel_activate['A' + str(each_cell)].value,
                                                excel_activate['B' + str(each_cell)].value,
                                                excel_activate['C' + str(each_cell)].value,
                                                excel_activate['D' + str(each_cell)].value,
                                                excel_activate['E' + str(each_cell)].value,
                                                excel_activate['F' + str(each_cell)].value,
                                                excel_activate['G' + str(each_cell)].value])
                        return updated_value

                    def back_to_login():
                        global amount
                        amount = 0
                        payorder.destroy()
                        login.deiconify()

                    each_cell_value = 3

                    def payment_function(each_cell):
                        excel_con = load_workbook('Orders.xlsx')
                        excel_activate = excel_con.active
    
                        payment_input = payent.get()
                        Found = False

                        if payment_input == "" or payment_input == "Enter Amount":
                            messagebox.showinfo("Notification", "Payment Required")
                        elif int(payment_input)<int(amount):
                            messagebox.showerror("Error","Amount is Not Enough")
                        else:
                            for each_cell in range(2, (excel_activate.max_row)+1):
                                if (first.get() == excel_activate['A'+str(each_cell)].value):
                                    Found = True
                                    break
                            else:
                                Found = False

                        if (Found == True):
                            cartlist_str = json.dumps(cartlist)

                            excel_activate['F'+str(each_cell)].value = cartlist_str
                            excel_activate['G'+str(each_cell)].value = amount
                            excel_con.save('Orders.xlsx')

                            name = excel_activate['A' + str(each_cell)].value
                            address = excel_activate['B' + str(each_cell)].value
                            users = excel_activate['C'+str(each_cell)].value
                            email = excel_activate['D' + str(each_cell)].value
                            mobileno = excel_activate['E' + str(each_cell)].value
                            order = excel_activate['F' + str(each_cell)].value
                            message = f'Thank You {name} !\nYour Order Will Be Delivered @ {address}\nYou Use The Username :\t{users}\nWe Will Send The Email Confirmation @ {email}\nWe Will Contact You In This #{mobileno}\n\nYour Order Are :\n{order} - PAID'
                            messagebox.showinfo("Notification", message)

                            cartlist.clear()
                            refresh_data(tree)

                            return
                    
                    def save_function():
                        excel_con = load_workbook('Orders.xlsx')
                        excel_activate = excel_con.active

                        if first.get() == "":
                            messagebox.showinfo("Notification", "FullName Required")
                        else:
                            lastrow = str(excel_activate.max_row + 1)

                            excel_activate['A'+str(lastrow)].value = first.get()
                            excel_activate['B'+str(lastrow)].value = last.get()
                            excel_activate['C'+str(lastrow)].value = mob.get()
                            excel_activate['D'+str(lastrow)].value = emails.get()
                            excel_activate['E'+str(lastrow)].value = usern.get()
                            excel_con.save('Orders.xlsx')

                            messagebox.showinfo("Notification", "Info Saved")

                            refresh_data(tree)


                    def back_to_menu():
                        global amount
                        amount = 0
                        payorder.destroy()
                        welc.deiconify()

                    payorderf1=Frame(payorder,height=150,width=1366)
                    payorderf1.pack(fill=BOTH,expand=1)

                    c=Canvas(payorderf1,height=150,width=1366)
                    c.pack()

                    logo=PhotoImage(file="logo.PNG")
                    c.image = logo
                    c.create_image(683,75,image=logo)

                    home=Button(payorderf1,text="Log Out",command=lambda:back_to_login(),bg="#0b1335",cursor="hand2",fg="white",bd=5,font=("default",16,'bold'))
                    home.place(x=1000,y=90)

                    localtime=time.asctime(time.localtime(time.time()))
                    c.create_text(1000,50,text=localtime,fill="white",font=("default",16))

                    payorderf2=Frame(payorder,height=618,width=1366)
                    payorderf2.pack(fill=BOTH,expand=1)

                    c=Canvas(payorderf2,height=618,width=1366)
                    c.pack()

                    logo1=PhotoImage(file="pizzamain.png")
                    c.image = logo1
                    c.create_image(683,309,image=logo1)
                    c.create_rectangle(30,30,1340,580,fill="#d3ede6",outline="white",width=6)

                    log=Label(payorderf2,text="Payment Order",fg="white",bg="#0b1335",width=20,font=("cooper black",27))
                    log.place(x=700,y=80)

                    info=Label(payorderf2,text="Information",fg="white",bg="#0b1335",width=20,font=("cooper black",27))
                    info.place(x=100,y=80)

                    lab1=Label(payorderf2,text="FullName",bg="#d3ede6",font=("cooper black",18))
                    lab1.place(x=100,y=140)

                    first=Entry(payorderf2,bg="white",width=15,font=("cooper black",18),bd=5)
                    first.place(x=300,y=140)

                    lab2=Label(payorderf2,text="Adress",bg="#d3ede6",font=("cooper black",18))
                    lab2.place(x=100,y=190)

                    last=Entry(payorderf2,bg="white",width=15,font=("cooper black",18),bd=5)
                    last.place(x=300,y=190)

                    lab3=Label(payorderf2,text="Username",bg="#d3ede6",font=("cooper black",18))
                    lab3.place(x=100,y=240)

                    usern=Entry(payorderf2,bg="white",width=15,font=("cooper black",18),bd=5)
                    usern.place(x=300,y=240)

                    lab6=Label(payorderf2,text="Email.",bg="#d3ede6",font=("cooper black",18))
                    lab6.place(x=100,y=290)

                    mob=Entry(payorderf2,bg="white",width=15,font=("cooper black",18),bd=5)
                    mob.place(x=300,y=290)

                    lab5=Label(payorderf2,text="Mobile No.",bg="#d3ede6",font=("cooper black",18))
                    lab5.place(x=100,y=340)

                    emails=Entry(payorderf2,bg="white",width=15,font=("cooper black",18),bd=5)
                    emails.place(x=300,y=340)

                    bc=Button(payorderf2,text="Back",cursor="hand2",fg="white",command=lambda:back_to_menu(),bg="#0b1335",font=("cooper black",18),bd=5)
                    bc.place(x=100,y=490)

                    sv=Button(payorderf2,text="Save",cursor="hand2",fg="white",bg="#0b1335",command=lambda:save_function(),font=("cooper black",18),bd=5)
                    sv.place(x=200,y=490)

                    rg=Button(payorderf2,text="Search Info",cursor="hand2",fg="white",bg="#0b1335",command=lambda:search(),font=("cooper black",18),bd=5)
                    rg.place(x=300,y=490)

                    cl=Button(payorderf2,text="Clear",cursor="hand2",fg="white",bg="#0b1335",command=lambda:clear(),font=("cooper black",18),bd=5)
                    cl.place(x=480,y=490)

                    c.create_text(940,150,text="Items\tSize\tQty\tPrice",font=("cooper black",18))
                    c.create_text(940,160,text="_______________________________________",font=("cooper black",18))
                    y=180
                    for i in cartlist:
                        y+=30
                        s=i[0]+"\t"+i[1]+"\t"+i[2]+"\t"+str(i[3])
                        c.create_text(940,y,text=s,font=("default",16))

                    amt=amount
                    text="Total : "+str(amt)
                    
                    tot=Label(payorderf2,text=text,bg="#f2da9d",width=12,font=("Cooper Black",22))
                    tot.place(x=700,y=400)

                    def on_entry_click(event):
                        if payent.get() == "Enter Amount":
                            payent.delete(0, tk.END)
                            payent.config(fg='black')
                    placeholder_text = "Enter Amount"

                    payment=Label(payorderf2,text="Payment",bg="#f2da9d",width=12,font=("Cooper Black",22))
                    payment.place(x=1050,y=400)

                    payent=Entry(payorderf2,bg="white",width=15,font=("cooper black",18),bd=5)
                    payent.place(x=1050,y=450)
                    payent.insert(0,placeholder_text)
                    payent.bind("<FocusIn>", on_entry_click)

                    rg=Button(payorderf2,text="PAY",cursor="hand2",fg="white",bg="#0b1335",command=lambda:payment_function(each_cell_value),font=("cooper black",18),bd=5)
                    rg.place(x=700,y=490)

                    def clear():
                        usern.delete(0,END)
                        first.delete(0,END)
                        last.delete(0,END)
                        emails.delete(0,END)
                        mob.delete(0,END)

                    def search():
                        excel_con = load_workbook('Orders.xlsx')
                        excel_activate = excel_con.active
                        
                        if first.get() == "":
                            messagebox.showinfo('NOTIFICATION', 'FullName REQUIRED')
                        else:
                            Found = False
                            for each_cell in range(2, (excel_activate.max_row) + 1):
                                if first.get() == excel_activate['A' + str(each_cell)].value:
                                    Found = True
                                    cell_address = str(each_cell)
                                    break

                            if Found:
                                address = excel_activate['B' + str(each_cell)].value
                                name = excel_activate['C' + str(each_cell)].value
                                mobile = excel_activate['D' + str(each_cell)].value
                                em = excel_activate['E' + str(each_cell)].value
                                
                                usern.insert(0,name)
                                last.insert(0,address)
                                emails.insert(0,em)
                                mob.insert(0,mobile)
                                
                                message = f'DATA EXIST IN {cell_address}'
                                messagebox.showinfo("FOUND",message)
                            else:
                                messagebox.showinfo("NOT FOUND", "NO INFO FOUND")
            
                    payorder.mainloop()
                
                ordf1=Frame(order,height=150,width=1366)
                ordf1.pack(fill=BOTH,expand=1)

                c=Canvas(ordf1,height=150,width=1366)
                c.pack()
                logo=PhotoImage(file="logo.PNG")
                c.image = logo
                c.create_image(683,75,image=logo)

                home=Button(ordf1,text="Log Out",command=lambda:back_to_login(),bg="#0b1335",cursor="hand2",fg="white",bd=5,font=("default",16,'bold'))
                home.place(x=1000,y=90)

                localtime=time.asctime(time.localtime(time.time()))
                c.create_text(1000,50,text=localtime,fill="white",font=("default",16))
                
                ordf2=Frame(order,height=618,width=1366)
                ordf2.pack(fill=BOTH,expand=1)

                c=Canvas(ordf2,height=618,width=1366)
                c.pack()
                logo1=PhotoImage(file="pizzamain.png")
                c.image = logo1
                c.create_image(683,309,image=logo1)

                log=Label(ordf2,text="YOUR ORDER",bg="#9db1f2",font=("Cooper Black",22))
                log.place(x=450,y=4)
                c.create_rectangle(250, 50, 800, 500,fill="#d3ede6",outline="white",width=6)

                amt=amount
                text="Total : "+str(amt)
                
                tot=Label(ordf2,text=text,bg="#f2da9d",width=12,font=("Cooper Black",22))
                tot.place(x=900,y=250)
                # if menulist =="deli":
                #     y=Address
                # if menulist=="pick":
                #     y=orderpay
                pay=Button(ordf2,text="Pay",command=lambda:orderpay(),bg="#0b1335",cursor="hand2",fg="white",bd=5,font=("default",18,'bold'))
                pay.place(x=900,y=300)

                exi=Button(ordf2,text="Add more",command=lambda:addmore_function(),bg="#0b1335",cursor="hand2",fg="white",bd=5,font=("default",18,'bold'))
                exi.place(x=1070,y=300)

                c.create_text(525,80,text="Items\tSize\tQty\tPrice",font=("cooper black",18))
                c.create_text(525,90,text="_______________________________________",font=("cooper black",18))
                y=100
                for i in cartlist:
                    y+=30
                    s=i[0]+"\t"+i[1]+"\t"+i[2]+"\t"+str(i[3])
                    c.create_text(525,y,text=s,font=("default",16))
                    
                
                order.mainloop()

            def vegpizza():
                veg=Toplevel()
                veg.title("Lavarez's Pizza")
                veg.geometry("1366x768")
                veg.resizable(False, False)
                menu.withdraw()

                def back_to_login():
                    veg.destroy()
                    login.deiconify()

                def addmore_function():
                    veg.destroy()
                    menu.deiconify()

                def order_function():
                    veg.destroy()
                    Orderde()

                vegf1=Frame(veg,height=150,width=1366)
                vegf1.pack(fill=BOTH,expand=1)

                c=Canvas(vegf1,height=150,width=1366)
                c.pack()

                logo=PhotoImage(file="logo.PNG")
                c.image = logo
                c.create_image(683,75,image=logo)

                home=Button(vegf1,text="Log Out",command=lambda:back_to_login(),bg="#0b1335",cursor="hand2",fg="white",bd=5,font=("default",16,'bold'))
                home.place(x=1000,y=90)

                localtime=time.asctime(time.localtime(time.time()))
                c.create_text(1000,50,text=localtime,fill="white",font=("default",16))

                vegf2=Frame(veg,height=618,width=1366)
                vegf2.pack(fill=BOTH,expand=1)
                
                c=Canvas(vegf2,height=618,width=1366)
                c.pack()
                logo1=PhotoImage(file="pizzamain.png")
                c.image = logo1
                c.create_image(683,309,image=logo1)

                log=Label(vegf2,text="VEG PIZZA",bg="#9db1f2",font=("Cooper Black",22))
                log.place(x=600,y=4)
                c.create_rectangle(400, 40, 966, 540,fill="#d3ede6",outline="white",width=6)

                q1=StringVar()
                q2=StringVar()
                q3=StringVar()
                q4=StringVar()
                q1.set("0")
                q2.set("0")
                q3.set("0")
                q4.set("0")

                # pizza 1
                c.create_rectangle(405, 50, 960, 170,width=2)
                delu=PhotoImage(file="deluxe.png")
                c.image = delu
                c.create_image(470,110,image=delu)
                c.create_text(650,80,text="Deluxe Veggie",fill="#000000",font=("Cooper Black",20))
                c.create_text(860,80,text="₱450/₱650/₱250",fill="#ff3838",font=("default",17,'bold'))

                #ch1=check(vegf2,100)
                v1=IntVar()
                C11=Radiobutton(vegf2,text = "Medium",value=10,variable=v1)
                C11.place(x=550,y=100)
                C12 = Radiobutton(vegf2, text = "Large",value = 20, variable =v1)
                C12.place(x=650,y=100)
                C13 = Radiobutton(vegf2, text = "Regular",value = 30, variable =v1)
                C13.place(x=750,y=100)
                C11.select()
                C11.deselect()    
                C11.invoke()

                c.create_text(590,150,text="Quantity : ",fill="#000000",font=("default",12))
                qty1=Entry(vegf2,textvariable=q1,bg="#aae2d7",font=("default",12),width=4,)
                qty1.place(x=650,y=140)

                add1=Button(vegf2,text="ADD",command=lambda:addch1(),bg="#0b1335",cursor="hand2",fg="white",bd=4,font=("default",12,'bold'))
                add1.place(x=850,y=120)
                def addch1():
                    if v1.get()==10:
                        ch1="Medium"
                        pric1=450
                    elif v1.get()==20:
                        ch1="Large"
                        pric1=650
                    else:
                        ch1="Regular"
                        pric1=250
                    addlist(["Deluxe Veggie",ch1,q1.get(),pric1*int(q1.get())])
                    
                #pizza 2
                c.create_rectangle(405, 170, 960, 290,width=2)
                vag=PhotoImage(file="extravaganza.png")
                c.image = vag
                c.create_image(470,230,image=vag)
                c.create_text(650,200,text="Veg Vaganza",fill="#000000",font=("Cooper Black",20))
                c.create_text(860,200,text="₱400/₱600/₱250",fill="#ff3838",font=("default",17,'bold'))

        ##        ch2=check(vegf2,220)
                v2=IntVar()
                C21=Radiobutton(vegf2,text = "Medium",value=10,variable=v2)
                C21.place(x=550,y=220)
                C22 = Radiobutton(vegf2, text = "Large",value = 20, variable =v2)
                C22.place(x=650,y=220)
                C23 = Radiobutton(vegf2, text = "Regular",value = 30, variable =v2)
                C23.place(x=750,y=220)
                C21.select()
                C21.deselect()    
                C21.invoke()

                c.create_text(590,270,text="Quantity : ",fill="#000000",font=("default",12))
                qty2=Entry(vegf2,textvariable=q2,bg="#aae2d7",font=("default",12),width=4,)
                qty2.place(x=650,y=260)

                add2=Button(vegf2,text="ADD",command=lambda:addch2(),bg="#0b1335",cursor="hand2",fg="white",bd=4,font=("default",12,'bold'))
                add2.place(x=850,y=240)
                def addch2():
                    if v2.get()==10:
                        ch2="Medium"
                        pric2=400
                    elif v2.get()==20:
                        ch2="Large"
                        pric2=600
                    else:
                        ch2="Regular"
                        pric2=250

                    addlist(["Veg Vaganza",ch2,q2.get(),pric2*int(q2.get())])

                #pizza 3
                c.create_rectangle(405, 290, 960, 410,width=2)
                pep=PhotoImage(file="5-pepper-veg-pizza.png")
                c.image = pep
                c.create_image(470,350,image=pep)
                c.create_text(650,320,text="5 Pepper",fill="#000000",font=("Cooper Black",20))
                c.create_text(860,320,text="₱385/₱550/₱225",fill="#ff3838",font=("default",17,'bold'))

                #ch3=check(vegf2,340)
                v3=IntVar()
                C31=Radiobutton(vegf2,text = "Medium",value=10,variable=v3)
                C31.place(x=550,y=340)
                C32 = Radiobutton(vegf2, text = "Large",value = 20, variable =v3)
                C32.place(x=650,y=340)
                C33 = Radiobutton(vegf2, text = "Regular",value = 30, variable =v3)
                C33.place(x=750,y=340)
                C31.select()
                C31.deselect()    
                C31.invoke()

                c.create_text(590,390,text="Quantity : ",fill="#000000",font=("default",12))
                qty3=Entry(vegf2,textvariable=q3,bg="#aae2d7",font=("default",12),width=4,)
                qty3.place(x=650,y=380)

                add3=Button(vegf2,text="ADD",command=lambda:addch3(),bg="#0b1335",cursor="hand2",fg="white",bd=4,font=("default",12,'bold'))
                add3.place(x=850,y=360)
                def addch3():
                    if v3.get()==10:
                        ch3="Medium"
                        pric3=385
                    elif v3.get()==20:
                        ch3="Large"
                        pric3=550
                    else:
                        ch3="Regular"
                        pric3=225
                    addlist(["5 Pepper     ",ch3,q3.get(),pric3*int(q3.get())])
                    
                #pizza 4
                c.create_rectangle(405, 410, 960, 530,width=2)
                mag=PhotoImage(file="Margherit.png")
                c.image = mag
                c.create_image(470,470,image=mag)
                c.create_text(650,440,text="Margherita",fill="#000000",font=("Cooper Black",20))
                c.create_text(860,440,text="₱195/₱385/₱99",fill="#ff3838",font=("default",17,'bold'))

                #ch4=check(vegf2,460)
                v4=IntVar()
                C41=Radiobutton(vegf2,text = "Medium",value=10,variable=v4)
                C41.place(x=550,y=460)
                C42 = Radiobutton(vegf2, text = "Large",value = 20, variable =v4)
                C42.place(x=650,y=460)
                C43 = Radiobutton(vegf2, text = "Regular",value = 30, variable =v4)
                C43.place(x=750,y=460)
                C41.select()
                C41.deselect()    
                C41.invoke()
                
                c.create_text(590,500,text="Quantity : ",fill="#000000",font=("default",12))
                qty4=Entry(vegf2,textvariable=q4,bg="#aae2d7",font=("default",12),width=4,)
                qty4.place(x=650,y=500)
                
                add4=Button(vegf2,text="ADD",command=lambda:addch4(),bg="#0b1335",cursor="hand2",fg="white",bd=4,font=("default",12,'bold'))
                add4.place(x=850,y=480)
                def addch4():
                    if v4.get()==10:
                        ch4="Medium"
                        pric4=195
                    elif v4.get()==20:
                        ch4="Large"
                        pric4=385
                    else:
                        ch4="Regular"
                        pric4=99
                    addlist(["Margherita  ",ch4,q4.get(),pric4*int(q4.get())])

                con=Button(vegf2,text="Confirm Order",command=lambda:order_function(),bg="#0b1335",cursor="hand2",fg="white",bd=5,font=("default",18,'bold'))
                con.place(x=1050,y=250)
                more=Button(vegf2,text="Add More..",command=lambda:addmore_function(),bg="#0b1335",cursor="hand2",fg="white",bd=5,font=("default",18,'bold'))
                more.place(x=1050,y=350)
                
                veg.mainloop()
            
            def nonvegpiz():
                nonveg=Toplevel()
                nonveg.title("Lavarez's Pizza")
                nonveg.geometry("1366x768")
                nonveg.resizable(False, False)
                menu.withdraw()

                def back_to_login():
                    nonveg.destroy()
                    login.deiconify()

                def addmore_function():
                    nonveg.destroy()
                    menu.deiconify()

                def order_function():
                    nonveg.destroy()
                    Orderde()


                nonvegf1=Frame(nonveg,height=150,width=1366)
                nonvegf1.pack(fill=BOTH,expand=1)

                c=Canvas(nonvegf1,height=150,width=1366)
                c.pack()
                logo=PhotoImage(file="logo.PNG")
                c.image=logo
                c.create_image(683,75,image=logo)

                home=Button(nonvegf1,text="Log Out",command=lambda:back_to_login(),bg="#0b1335",cursor="hand2",fg="white",bd=5,font=("default",16,'bold'))
                home.place(x=1000,y=90)

                localtime=time.asctime(time.localtime(time.time()))
                c.create_text(1000,50,text=localtime,fill="white",font=("default",16))

                nonvegf2=Frame(nonveg,height=618,width=1366)
                nonvegf2.pack(fill=BOTH,expand=1)
        
                c=Canvas(nonvegf2,height=618,width=1366)
                c.pack()

                logo1=PhotoImage(file="pizzamain.png")
                c.image= logo1
                c.create_image(683,309,image=logo1)

                log=Label(nonvegf2,text="NON-VEG PIZZA",bg="#9db1f2",font=("Cooper Black",22))
                log.place(x=580,y=4)
                c.create_rectangle(400, 40, 966, 540,fill="#d3ede6",outline="white",width=6)

                q5=StringVar()
                q6=StringVar()
                q7=StringVar()
                q8=StringVar()
                q5.set("0")
                q6.set("0")
                q7.set("0")
                q8.set("0")

                # pizza 1
                c.create_rectangle(405, 50, 960, 170,width=2)
                delu=PhotoImage(file="Non-Veg_Supreme.png")
                c.image= delu
                c.create_image(470,110,image=delu)
                c.create_text(650,80,text="Non-Veg Supreme",fill="#000000",font=("Cooper Black",20))
                c.create_text(860,80,text="₱450/₱650/₱250",fill="#ff3838",font=("default",17,'bold'))

                #ch5=check(nonvegf2,100)
                v5=IntVar()
                C51=Radiobutton(nonvegf2,text = "Medium",value=10,variable=v5)
                C51.place(x=550,y=100)
                C52 = Radiobutton(nonvegf2, text = "Large",value = 20, variable =v5)
                C52.place(x=650,y=100)
                C53 = Radiobutton(nonvegf2, text = "Regular",value = 30, variable =v5)
                C53.place(x=750,y=100)
                C51.select()
                C51.deselect()    
                C51.invoke()

                qty5=Entry(nonvegf2,textvariable=q5,bg="#aae2d7",font=("default",12),width=4,)
                qty5.place(x=650,y=140)
                
                add5=Button(nonvegf2,text="ADD",command=lambda:addch5(),bg="#0b1335",cursor="hand2",fg="white",bd=4,font=("default",12,'bold'))
                add5.place(x=850,y=120)
                def addch5():
                    if v5.get()==10:
                        ch5="Medium"
                        pric5=450
                    elif v5.get()==20:
                        ch5="Large"
                        pric5=650
                    else:
                        ch5="Regular"
                        pric5=250
                    addlist(["Non-Veg Supreme",ch5,q5.get(),pric5*int(q5.get())])

                #pizza 2
                c.create_rectangle(405, 170, 960, 290,width=2)
                vag=PhotoImage(file="nonChicken_Tikka.png")
                c.image= vag
                c.create_image(470,230,image=vag)
                c.create_text(650,200,text="Chicken Tikka",fill="#000000",font=("Cooper Black",20))
                c.create_text(860,200,text="₱400/₱600/₱250",fill="#ff3838",font=("default",17,'bold'))

                #ch6=check(nonvegf2,220)
                v6=IntVar()
                C61=Radiobutton(nonvegf2,text = "Medium",value=10,variable=v6)
                C61.place(x=550,y=220)
                C62 = Radiobutton(nonvegf2, text = "Large",value = 20, variable =v6)
                C62.place(x=650,y=220)
                C63 = Radiobutton(nonvegf2, text = "Regular",value = 30, variable =v6)
                C63.place(x=750,y=220)
                C61.select()
                C61.deselect()    
                C61.invoke()

                c.create_text(590,270,text="Quantity : ",fill="#000000",font=("default",12))
                qty6=Entry(nonvegf2,textvariable=q6,bg="#aae2d7",font=("default",12),width=4,)
                qty6.place(x=650,y=260)
                
                add6=Button(nonvegf2,text="ADD",command=lambda:addch6(),bg="#0b1335",cursor="hand2",fg="white",bd=4,font=("default",12,'bold'))
                add6.place(x=850,y=240)
                def addch6():
                    if v6.get()==10:
                        ch6="Medium"
                        pric6=400
                    elif v6.get()==20:
                        ch6="Large"
                        pric6=600
                    else:
                        ch6="Regular"
                        pric6=250
                    addlist(["Chicken Tikka",ch6,q6.get(),pric6*int(q6.get())])

                #pizza 3
                c.create_rectangle(405, 290, 960, 410,width=2)
                pep=PhotoImage(file="non-Chicken_Sausage.png")
                c.image= pep
                c.create_image(470,350,image=pep)
                c.create_text(650,320,text="Chicken Sausage",fill="#000000",font=("Cooper Black",20))
                c.create_text(860,320,text="₱385/₱550/₱225",fill="#ff3838",font=("default",17,'bold'))

                #ch7=check(nonvegf2,340)
                v7=IntVar()
                C71=Radiobutton(nonvegf2,text = "Medium",value=10,variable=v7)
                C71.place(x=550,y=340)
                C72 = Radiobutton(nonvegf2, text = "Large",value = 20, variable =v7)
                C72.place(x=650,y=340)
                C73 = Radiobutton(nonvegf2, text = "Regular",value = 30, variable =v7)
                C73.place(x=750,y=340)
                C71.select()
                C71.deselect()    
                C71.invoke()

                c.create_text(590,390,text="Quantity : ",fill="#000000",font=("default",12))
                qty7=Entry(nonvegf2,textvariable=q7,bg="#aae2d7",font=("default",12),width=4,)
                qty7.place(x=650,y=380)
                
                add7=Button(nonvegf2,text="ADD",command=lambda:addch7(),bg="#0b1335",cursor="hand2",fg="white",bd=4,font=("default",12,'bold'))
                add7.place(x=850,y=360)
                def addch7():
                    if v7.get()==10:
                        ch7="Medium"
                        pric7=385
                    elif v7.get()==20:
                        ch7="Large"
                        pric7=550
                    else:
                        ch7="Regular"
                        pric7=225
                    addlist(["Chicken Sausage",ch7,q7.get(),pric7*int(q7.get())])

                #pizza 4
                c.create_rectangle(405, 410, 960, 530,width=2)
                mag=PhotoImage(file="no-LoadedL.png")
                c.image= mag
                c.create_image(470,470,image=mag)
                c.create_text(650,440,text="Chicken Peri",fill="#000000",font=("Cooper Black",20))
                c.create_text(860,440,text="₱195/₱385/₱99",fill="#ff3838",font=("default",17,'bold'))

                #ch8=check(nonvegf2,460)
                v8=IntVar()
                C81=Radiobutton(nonvegf2,text = "Medium",value=10,variable=v8)
                C81.place(x=550,y=460)
                C82 = Radiobutton(nonvegf2, text = "Large",value = 20, variable =v8)
                C82.place(x=650,y=460)
                C83 = Radiobutton(nonvegf2, text = "Regular",value = 30, variable =v8)
                C83.place(x=750,y=460)
                C81.select()
                C81.deselect()    
                C81.invoke()

                c.create_text(590,500,text="Quantity : ",fill="#000000",font=("default",12))
                qty8=Entry(nonvegf2,textvariable=q8,bg="#aae2d7",font=("default",12),width=4,)
                qty8.place(x=650,y=500)

                add8=Button(nonvegf2,text="ADD",command=lambda:addch8(),bg="#0b1335",cursor="hand2",fg="white",bd=4,font=("default",12,'bold'))
                add8.place(x=850,y=480)
                def addch8():
                    if v8.get()==10:
                        ch8="Medium"
                        pric8=195
                    elif v8.get()==20:
                        ch8="Large"
                        pric8=385
                    else:
                        ch8="Regular"
                        pric8=99
                    addlist(["Chicken Peri",ch8,q8.get(),pric8*int(q8.get())])

                con=Button(nonvegf2,text="Confirm Order",command=lambda:order_function(),bg="#0b1335",cursor="hand2",fg="white",bd=5,font=("default",18,'bold'))
                con.place(x=1050,y=250)

                more=Button(nonvegf2,text="Add More..",command=lambda:addmore_function(),bg="#0b1335",cursor="hand2",fg="white",bd=5,font=("default",18,'bold'))
                more.place(x=1050,y=350)
                
        
                nonveg.mainloop()

            def SpecialChi():
                chicken=Toplevel()
                chicken.title("Lavarez's Pizza")
                chicken.geometry("1366x768")
                chicken.resizable(False, False)
                menu.withdraw()

                def back_to_login():
                    chicken.destroy()
                    login.deiconify()

                def addmore_function():
                    chicken.destroy()
                    menu.deiconify()

                def order_function():
                    chicken.destroy()
                    Orderde()

                spef1=Frame(chicken,height=150,width=1366)
                spef1.pack(fill=BOTH,expand=1)

                c=Canvas(spef1,height=150,width=1366)
                c.pack()

                logo=PhotoImage(file="logo.PNG")
                c.create_image(683,75,image=logo)

                home=Button(spef1,text="Log Out",command=lambda:back_to_login(),bg="#0b1335",cursor="hand2",fg="white",bd=5,font=("default",16,'bold'))
                home.place(x=1000,y=90)

                localtime=time.asctime(time.localtime(time.time()))
                c.create_text(1000,50,text=localtime,fill="white",font=("default",16))
                
                spef2=Frame(chicken,height=618,width=1366)
                
                c=Canvas(spef2,height=618,width=1366)
                c.pack()
                logo1=PhotoImage(file="pizzamain.png")
                c.create_image(683,309,image=logo1)

                log=Label(spef2,text="SPECIALTY CHICKEN",bg="#9db1f2",font=("Cooper Black",22))
                log.place(x=540,y=4)

                c.create_rectangle(400, 40, 966, 420,fill="#d3ede6",outline="white",width=6)
                q9=StringVar()
                q10=StringVar()
                q11=StringVar()
                q9.set("0")
                q10.set("0")
                q11.set("0")

                # chicken 1
                c.create_rectangle(405, 50, 960, 170,width=2)
                delu=PhotoImage(file="roasted.png")
                c.create_image(470,110,image=delu)
                c.create_text(650,80,text="Roasted Chicken",fill="#000000",font=("Cooper Black",20))
                c.create_text(875,80,text="₱250/₱450/₱150",fill="#ff3838",font=("default",14,'bold'))
                c.create_text(590,150,text="Quantity : ",fill="#000000",font=("default",12))

                v5=IntVar()
                C51=Radiobutton(spef2,text = "Medium",value=10,variable=v5)
                C51.place(x=550,y=100)
                C52 = Radiobutton(spef2, text = "Large",value = 20, variable =v5)
                C52.place(x=650,y=100)
                C53 = Radiobutton(spef2, text = "Regular",value = 30, variable =v5)
                C53.place(x=750,y=100)
                C51.select()
                C51.deselect()    
                C51.invoke()

                qty9=Entry(spef2,textvariable=q9,bg="#aae2d7",font=("default",12),width=4,)
                qty9.place(x=650,y=140)

                add9=Button(spef2,text="ADD",command=lambda:add_func(),bg="#0b1335",cursor="hand2",fg="white",bd=4,font=("default",12,'bold'))
                add9.place(x=850,y=120)

                def add_func():
                    if v5.get()==10:
                        ch5="Medium"
                        pric5=250
                    elif v5.get()==20:
                        ch5="Large"
                        pric5=450
                    else:
                        ch5="Regular"
                        pric5=150
                    addlist(["Roasted Chicken",ch5,q9.get(),pric5*int(q9.get())])

                #chicken 2
                c.create_rectangle(405, 170, 960, 290,width=2)
                vag=PhotoImage(file="chicken-meatballs.jpg")
                c.create_image(470,230,image=vag)
                c.create_text(650,200,text="Chicken Meatballs",fill="#000000",font=("Cooper Black",20))
                c.create_text(875,200,text="₱149/₱199/₱99",fill="#ff3838",font=("default",14,'bold'))
                c.create_text(590,270,text="Quantity : ",fill="#000000",font=("default",12))

                v6=IntVar()
                C61=Radiobutton(spef2,text = "M(8pcs)",value=10,variable=v6)
                C61.place(x=550,y=220)
                C62 = Radiobutton(spef2, text = "L(12pcs)",value = 20, variable =v6)
                C62.place(x=650,y=220)
                C63 = Radiobutton(spef2, text = "R(4pcs)",value = 30, variable =v6)
                C63.place(x=750,y=220)
                C61.select()
                C61.deselect()    
                C61.invoke()

                qty10=Entry(spef2,textvariable=q10,bg="#aae2d7",font=("default",12),width=4,)
                qty10.place(x=650,y=260)

                add10=Button(spef2,text="ADD",command=lambda:add_func2(),bg="#0b1335",cursor="hand2",fg="white",bd=4,font=("default",12,'bold'))
                add10.place(x=850,y=240)

                def add_func2():
                    if v6.get()==10:
                        ch6="M(8pcs)"
                        pric6=149
                    elif v6.get()==20:
                        ch6="L(12pcs)"
                        pric6=199
                    else:
                        ch6="R(4pcs)"
                        pric6=99
                    addlist(["Chicken Meatballs",ch6,q10.get(),pric6*int(q10.get())])
                
                #chicken 3
                c.create_rectangle(405, 290, 960, 410,width=2)
                pep=PhotoImage(file="Boneless-Chicken-wings-192x192.png")
                c.create_image(470,350,image=pep)
                c.create_text(650,320,text="Boneless Chicken",fill="#000000",font=("Cooper Black",20))
                c.create_text(875,320,text="₱199/₱239/₱139",fill="#ff3838",font=("default",14,'bold'))
                c.create_text(590,390,text="Quantity : ",fill="#000000",font=("default",12))

                v7=IntVar()
                C71=Radiobutton(spef2,text = "M(8pcs)",value=10,variable=v7)
                C71.place(x=550,y=340)
                C72 = Radiobutton(spef2, text = "L(12pcs)",value = 20, variable =v7)
                C72.place(x=650,y=340)
                C73 = Radiobutton(spef2, text = "R(4pcs)",value = 30, variable =v7)
                C73.place(x=750,y=340)
                C71.select()
                C71.deselect()    
                C71.invoke()

                qty11=Entry(spef2,textvariable=q11,bg="#aae2d7",font=("default",12),width=4,)
                qty11.place(x=650,y=380)

                add11=Button(spef2,text="ADD",command=lambda:add_func3(),bg="#0b1335",cursor="hand2",fg="white",bd=4,font=("default",12,'bold'))
                add11.place(x=850,y=360)

                def add_func3():
                    if v7.get()==10:
                        ch7="M(8pcs)"
                        pric7=199
                    elif v7.get()==20:
                        ch7="L(12pcs)"
                        pric7=239
                    else:
                        ch7="R(4pcs)"
                        pric7=139
                    addlist(["Boneless Chiken",ch7,q11.get(),pric7*int(q11.get())])

                con=Button(spef2,text="Confirm Order",command=lambda:order_function(),bg="#0b1335",cursor="hand2",fg="white",bd=5,font=("default",18,'bold'))
                con.place(x=600,y=430)

                more=Button(spef2,text="Add More..",command=lambda:addmore_function(),bg="#0b1335",cursor="hand2",fg="white",bd=5,font=("default",16,'bold'))
                more.place(x=630,y=500)
                spef2.pack(fill=BOTH,expand=1)
                chicken.mainloop()

            def sidebev():
                side=Toplevel()
                side.title("Lavarez's Pizza")
                side.geometry("1366x768")
                side.resizable(False, False)
                menu.withdraw()

                def back_to_login():
                    side.destroy()
                    login.deiconify()

                def addmore_function():
                    side.destroy()
                    menu.deiconify()

                def order_function():
                    side.destroy()
                    Orderde()

                sidef1=Frame(side,height=150,width=1366)
                sidef1.pack(fill=BOTH,expand=1)

                c=Canvas(sidef1,height=150,width=1366)
                c.pack()
                logo=PhotoImage(file="logo.PNG")
                c.create_image(683,75,image=logo)

                home=Button(sidef1,text="Log Out",command=lambda:back_to_login(),bg="#0b1335",cursor="hand2",fg="white",bd=5,font=("default",16,'bold'))
                home.place(x=1000,y=90)

                localtime=time.asctime(time.localtime(time.time()))
                c.create_text(1000,50,text=localtime,fill="white",font=("default",16))
                
                sidef2=Frame(side,height=618,width=1366)
                sidef2.pack(fill=BOTH,expand=1)
                
                c=Canvas(sidef2,height=618,width=1366)
                c.pack()
                logo1=PhotoImage(file="pizzamain.png")
                c.create_image(683,309,image=logo1)

                log=Label(sidef2,text="SIDES & BEVERAGES",bg="#9db1f2",font=("Cooper Black",22))
                log.place(x=520,y=4)
                c.create_rectangle(400, 40, 966, 420,fill="#d3ede6",outline="white",width=6)

                q12=StringVar()
                q13=StringVar()
                q14=StringVar()
                q12.set("0")
                q13.set("0")
                q14.set("0")

                # side beverage 1

                c.create_rectangle(405, 50, 960, 170,width=2)
                delu=PhotoImage(file="coke.png")
                c.create_image(470,110,image=delu)

                c.create_text(650,80,text="Coke Mobile",fill="#000000",font=("Cooper Black",20))
                c.create_text(875,80,text="₱45/₱60/₱30",fill="#ff3838",font=("default",14,'bold'))
                c.create_text(590,140,text="Quantity : ",fill="#000000",font=("default",12))

                v5=IntVar()
                C51=Radiobutton(sidef2,text = "Medium",value=10,variable=v5)
                C51.place(x=550,y=100)
                C52 = Radiobutton(sidef2, text = "Large",value = 20, variable =v5)
                C52.place(x=650,y=100)
                C53 = Radiobutton(sidef2, text = "Regular",value = 30, variable =v5)
                C53.place(x=750,y=100)
                C51.select()
                C51.deselect()    
                C51.invoke()

                qty12=Entry(sidef2,textvariable=q12,bg="#aae2d7",font=("default",12),width=4,)
                qty12.place(x=650,y=130)

                add12=Button(sidef2,text="ADD",command=lambda:addch5(),bg="#0b1335",cursor="hand2",fg="white",bd=4,font=("default",12,'bold'))
                add12.place(x=850,y=120)
                def addch5():
                    if v5.get()==10:
                        ch5="Medium"
                        pric5=45
                    elif v5.get()==20:
                        ch5="Large"
                        pric5=60
                    else:
                        ch5="Regular"
                        pric5=30
                    addlist(["Coke Mobile",ch5,q12.get(),pric5*int(q12.get())])

                # side beverage 2

                c.create_rectangle(405, 170, 960, 290,width=2)
                vag=PhotoImage(file="burger.png")
                c.create_image(470,230,image=vag)
                
                c.create_text(650,200,text="Burger Pizza",fill="#000000",font=("Cooper Black",20))
                c.create_text(875,200,text="₱199/₱139",fill="#ff3838",font=("default",14,'bold'))
                c.create_text(590,270,text="Quantity : ",fill="#000000",font=("default",12))
                
                v6=IntVar()
                C61=Radiobutton(sidef2,text = "Quarter",value=10,variable=v6)
                C61.place(x=550,y=220)
                C62 = Radiobutton(sidef2, text = "Regular",value = 20, variable =v6)
                C62.place(x=700,y=220)
                C61.select()
                C61.deselect()    
                C61.invoke()

                qty13=Entry(sidef2,textvariable=q13,bg="#aae2d7",font=("default",12),width=4,)
                qty13.place(x=650,y=260)

                add13=Button(sidef2,text="ADD",command=lambda:addch6(),bg="#0b1335",cursor="hand2",fg="white",bd=4,font=("default",12,'bold'))
                add13.place(x=850,y=240)

                def addch6():
                    if v6.get()==10:
                        ch6="Quarter"
                        pric6=199
                    else:
                        ch6="Regular"
                        pric6=139
                    addlist(["Burger Pizza",ch6,q13.get(),pric6*int(q13.get())])


                # side beverage 3

                c.create_rectangle(405, 290, 960, 410,width=2)
                pep=PhotoImage(file="white.png")
                c.create_image(470,350,image=pep)

                c.create_text(650,320,text="White Pasta",fill="#000000",font=("Cooper Black",20))
                c.create_text(875,320,text="₱169/₱229/₱229",fill="#ff3838",font=("default",14,'bold'))
                c.create_text(590,390,text="Quantity : ",fill="#000000",font=("default",12))

                v7=IntVar()
                C71=Radiobutton(sidef2,text = "Medium",value=10,variable=v7)
                C71.place(x=550,y=340)
                C72 = Radiobutton(sidef2, text = "Large",value = 20, variable =v7)
                C72.place(x=650,y=340)
                C73 = Radiobutton(sidef2, text = "Regular",value = 30, variable =v7)
                C73.place(x=750,y=340)
                C71.select()
                C71.deselect()    
                C71.invoke()

                qty14=Entry(sidef2,textvariable=q14,bg="#aae2d7",font=("default",12),width=4,)
                qty14.place(x=650,y=380)

                add14=Button(sidef2,text="ADD",command=lambda:addch7(),bg="#0b1335",cursor="hand2",fg="white",bd=4,font=("default",12,'bold'))
                add14.place(x=850,y=360)

                def addch7():
                    if v7.get()==10:
                        ch7="Medium"
                        pric7=169
                    elif v7.get()==20:
                        ch7="Large"
                        pric7=229
                    else:
                        ch7="Regular"
                        pric7=119
                    addlist(["White Pasta",ch7,q14.get(),pric7*int(q14.get())])

                con=Button(sidef2,text="Confirm Order",command=lambda:order_function(),bg="#0b1335",cursor="hand2",fg="white",bd=5,font=("default",18,'bold'))
                con.place(x=600,y=430)

                more=Button(sidef2,text="Add More..",command=lambda:addmore_function(),bg="#0b1335",cursor="hand2",fg="white",bd=5,font=("default",16,'bold'))
                more.place(x=630,y=500)
                
                side.mainloop()

            menu.mainloop()
        
        welc.mainloop()

main.mainloop()